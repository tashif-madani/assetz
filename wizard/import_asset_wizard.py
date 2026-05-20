import base64
import io
import logging

from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    _logger.warning("openpyxl not installed. Excel import will be disabled.")


class ImportAssetWizard(models.TransientModel):
    """Wizard for importing assets from external ERP or CSV/Excel files."""

    _name = "assetz.import.asset.wizard"
    _description = "Import Asset Wizard"

    name = fields.Char(string="Import Name", default="Asset Import")
    file_data = fields.Binary(
        string="File",
        required=True,
        help="Upload a CSV or Excel file with asset data",
    )
    file_name = fields.Char(string="File Name")
    file_type = fields.Selection(
        selection=[
            ("csv", "CSV"),
            ("xlsx", "Excel (XLSX)"),
        ],
        string="File Type",
        compute="_compute_file_type",
        store=True,
    )

    # Import options
    skip_first_row = fields.Boolean(
        string="Skip Header Row",
        default=True,
        help="Skip the first row (header) of the file",
    )
    update_existing = fields.Boolean(
        string="Update Existing Assets",
        default=False,
        help="If checked, existing assets (matched by code or serial number) will be updated",
    )
    duplicate_check_field = fields.Selection(
        selection=[
            ("code", "Asset Code"),
            ("serial_number", "Serial Number"),
            ("both", "Both"),
        ],
        string="Duplicate Check By",
        default="code",
        help="Field to use for detecting duplicate assets",
    )

    # Column mapping
    col_name = fields.Integer(string="Name Column", default=1)
    col_code = fields.Integer(string="Code Column", default=0)
    col_serial_number = fields.Integer(string="Serial Number Column", default=2)
    col_category = fields.Integer(string="Category Column", default=3)
    col_location = fields.Integer(string="Location Column", default=4)
    col_acquisition_date = fields.Integer(string="Acquisition Date Column", default=5)
    col_acquisition_cost = fields.Integer(string="Acquisition Cost Column", default=6)
    col_manufacturer = fields.Integer(string="Manufacturer Column", default=7)
    col_model = fields.Integer(string="Model Column", default=8)
    col_description = fields.Integer(string="Description Column", default=9)

    # Results
    import_result = fields.Text(
        string="Import Result",
        readonly=True,
    )
    state = fields.Selection(
        selection=[
            ("draft", "Draft"),
            ("preview", "Preview"),
            ("done", "Done"),
        ],
        string="Status",
        default="draft",
    )

    # Preview data
    preview_line_ids = fields.One2many(
        comodel_name="assetz.import.asset.preview",
        inverse_name="wizard_id",
        string="Preview Lines",
    )

    @api.depends("file_name")
    def _compute_file_type(self):
        """Determine file type from extension."""
        for wizard in self:
            if wizard.file_name:
                if wizard.file_name.lower().endswith(".csv"):
                    wizard.file_type = "csv"
                elif wizard.file_name.lower().endswith((".xlsx", ".xls")):
                    wizard.file_type = "xlsx"
                else:
                    wizard.file_type = False
            else:
                wizard.file_type = False

    def action_preview(self):
        """Preview the import data."""
        self.ensure_one()

        if not self.file_data:
            raise UserError("Please upload a file first.")

        # Clear existing preview lines
        self.preview_line_ids.unlink()

        # Parse file
        rows = self._parse_file()

        # Create preview lines
        start_row = 1 if self.skip_first_row else 0
        preview_vals = []

        for idx, row in enumerate(rows[start_row:], start=start_row + 1):
            if idx > 20:  # Limit preview to 20 rows
                break

            preview_vals.append({
                "wizard_id": self.id,
                "row_number": idx,
                "name": self._get_cell_value(row, self.col_name),
                "code": self._get_cell_value(row, self.col_code),
                "serial_number": self._get_cell_value(row, self.col_serial_number),
                "category": self._get_cell_value(row, self.col_category),
                "location": self._get_cell_value(row, self.col_location),
                "acquisition_date": self._get_cell_value(row, self.col_acquisition_date),
                "acquisition_cost": self._get_cell_value(row, self.col_acquisition_cost),
                "status": "pending",
            })

        self.env["assetz.import.asset.preview"].create(preview_vals)
        self.state = "preview"

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_import(self):
        """Execute the import."""
        self.ensure_one()

        if not self.file_data:
            raise UserError("Please upload a file first.")

        rows = self._parse_file()
        start_row = 1 if self.skip_first_row else 0

        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0
        errors = []

        Asset = self.env["assetz.asset"]
        Category = self.env["assetz.category"]
        Location = self.env["assetz.location"]

        for idx, row in enumerate(rows[start_row:], start=start_row + 1):
            try:
                name = self._get_cell_value(row, self.col_name)
                code = self._get_cell_value(row, self.col_code)
                serial_number = self._get_cell_value(row, self.col_serial_number)

                if not name:
                    skipped_count += 1
                    continue

                # Check for duplicates
                existing = False
                if self.duplicate_check_field in ("code", "both") and code:
                    existing = Asset.search([("code", "=", code)], limit=1)
                if not existing and self.duplicate_check_field in ("serial_number", "both") and serial_number:
                    existing = Asset.search([("serial_number", "=", serial_number)], limit=1)

                # Prepare values
                vals = {
                    "name": name,
                    "serial_number": serial_number,
                    "manufacturer": self._get_cell_value(row, self.col_manufacturer),
                    "model": self._get_cell_value(row, self.col_model),
                    "description": self._get_cell_value(row, self.col_description),
                }

                # Handle category
                category_name = self._get_cell_value(row, self.col_category)
                if category_name:
                    category = Category.search([("name", "=ilike", category_name)], limit=1)
                    if not category:
                        category = Category.create({"name": category_name})
                    vals["category_id"] = category.id

                # Handle location
                location_name = self._get_cell_value(row, self.col_location)
                if location_name:
                    location = Location.search([("name", "=ilike", location_name)], limit=1)
                    if not location:
                        location = Location.create({"name": location_name})
                    vals["location_id"] = location.id

                # Handle acquisition date
                acq_date = self._get_cell_value(row, self.col_acquisition_date)
                if acq_date:
                    try:
                        from datetime import datetime
                        if isinstance(acq_date, str):
                            # Try common date formats
                            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                                try:
                                    vals["acquisition_date"] = datetime.strptime(acq_date, fmt).date()
                                    break
                                except ValueError:
                                    continue
                        else:
                            vals["acquisition_date"] = acq_date
                    except Exception:
                        pass

                # Handle acquisition cost
                acq_cost = self._get_cell_value(row, self.col_acquisition_cost)
                if acq_cost:
                    try:
                        if isinstance(acq_cost, str):
                            acq_cost = acq_cost.replace(",", "").replace("$", "").replace("€", "")
                        vals["acquisition_cost"] = float(acq_cost)
                    except (ValueError, TypeError):
                        pass

                if existing and self.update_existing:
                    existing.write(vals)
                    updated_count += 1
                elif existing:
                    skipped_count += 1
                else:
                    if code:
                        vals["code"] = code
                    Asset.create(vals)
                    created_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f"Row {idx}: {str(e)}")
                _logger.error(f"Import error at row {idx}: {e}")

        # Generate result summary
        result = f"""Import Complete!

Created: {created_count} assets
Updated: {updated_count} assets
Skipped: {skipped_count} rows
Errors: {error_count}
"""
        if errors:
            result += "\nErrors:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                result += f"\n... and {len(errors) - 10} more errors"

        self.write({
            "import_result": result,
            "state": "done",
        })

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def _parse_file(self):
        """Parse the uploaded file and return rows."""
        if not self.file_data:
            return []

        file_content = base64.b64decode(self.file_data)

        if self.file_type == "csv":
            return self._parse_csv(file_content)
        elif self.file_type == "xlsx":
            return self._parse_excel(file_content)
        else:
            raise UserError("Unsupported file type. Please upload a CSV or Excel file.")

    def _parse_csv(self, file_content):
        """Parse CSV file content."""
        import csv
        content = file_content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(content))
        return list(reader)

    def _parse_excel(self, file_content):
        """Parse Excel file content."""
        if not HAS_OPENPYXL:
            raise UserError("openpyxl library is required for Excel import. Please install it.")

        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows

    def _get_cell_value(self, row, col_index):
        """Get cell value from row by column index."""
        if col_index < 0 or col_index >= len(row):
            return ""
        value = row[col_index]
        if value is None:
            return ""
        return str(value).strip() if isinstance(value, str) else value

    def action_download_template(self):
        """Download a sample import template."""
        if not HAS_OPENPYXL:
            raise UserError("openpyxl library is required to generate template.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assets"

        # Header row
        headers = [
            "Code", "Name", "Serial Number", "Category", "Location",
            "Acquisition Date", "Acquisition Cost", "Manufacturer", "Model", "Description"
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Sample data row
        sample = [
            "AST00001", "Dell Laptop XPS 15", "SN123456", "IT Equipment", "Main Office",
            "2024-01-15", "1500.00", "Dell", "XPS 15 9530", "Employee laptop"
        ]
        for col, value in enumerate(sample, start=1):
            ws.cell(row=2, column=col, value=value)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env["ir.attachment"].create({
            "name": "asset_import_template.xlsx",
            "datas": base64.b64encode(output.getvalue()),
            "type": "binary",
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }


class ImportAssetPreview(models.TransientModel):
    """Preview line for asset import."""

    _name = "assetz.import.asset.preview"
    _description = "Import Asset Preview Line"

    wizard_id = fields.Many2one(
        comodel_name="assetz.import.asset.wizard",
        string="Wizard",
        required=True,
        ondelete="cascade",
    )
    row_number = fields.Integer(string="Row #")
    name = fields.Char(string="Name")
    code = fields.Char(string="Code")
    serial_number = fields.Char(string="Serial Number")
    category = fields.Char(string="Category")
    location = fields.Char(string="Location")
    acquisition_date = fields.Char(string="Acquisition Date")
    acquisition_cost = fields.Char(string="Cost")
    status = fields.Selection(
        selection=[
            ("pending", "Pending"),
            ("duplicate", "Duplicate"),
            ("error", "Error"),
        ],
        string="Status",
        default="pending",
    )
    notes = fields.Char(string="Notes")
